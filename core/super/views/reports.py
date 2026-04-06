"""
Vista de reportes y estadísticas de ventas con exportación a Excel.
"""

import io
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Sum, Avg, Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.generic import TemplateView
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

from core.super.models import Sale, SaleDetail, Product, Category, Customer, Seller


# Helpers

def _parse_date(value, fallback):
    """Convierte string YYYY-MM-DD a date, devuelve fallback si falla."""
    if not value:
        return fallback
    try:
        return date.fromisoformat(value)
    except ValueError:
        return fallback


def _apply_filters(request, force_all=False):
    """Lee los parámetros GET y devuelve el queryset de Sale filtrado.
    Si force_all=True, ignora los filtros y devuelve todo el histórico."""
    today = date.today()

    if force_all:
        date_from = date(2000, 1, 1)
        date_to   = today
        qs = Sale.objects.all()
        return qs, date_from, date_to

    date_from = _parse_date(request.GET.get("date_from"), today - timedelta(days=30))
    date_to   = _parse_date(request.GET.get("date_to"),   today)
    payment     = request.GET.get("payment", "")
    seller_id   = request.GET.get("seller", "")
    customer_id = request.GET.get("customer", "")
    min_total   = request.GET.get("min_total", "")
    max_total   = request.GET.get("max_total", "")

    qs = Sale.objects.filter(
        sale_date__date__gte=date_from,
        sale_date__date__lte=date_to,
    )
    if payment:
        qs = qs.filter(payment__name__icontains=payment)
    if seller_id:
        qs = qs.filter(seller_id=seller_id)
    if customer_id:
        qs = qs.filter(customer_id=customer_id)
    if min_total:
        qs = qs.filter(total__gte=Decimal(min_total))
    if max_total:
        qs = qs.filter(total__lte=Decimal(max_total))

    return qs, date_from, date_to


# Vista principal

class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = "super/reports/reports.html"
    login_url = "/security/login/"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request

        qs, date_from, date_to = _apply_filters(request)

        # KPIs generales
        agg = qs.aggregate(
            total_revenue=Sum("total"),
            total_sales=Count("id_sale"),
            total_discount=Sum("discount"),
        )
        total_revenue = agg["total_revenue"] or Decimal("0.00")
        total_sales   = agg["total_sales"]   or 0
        total_discount= agg["total_discount"]or Decimal("0.00")
        avg_ticket    = (total_revenue / total_sales) if total_sales else Decimal("0.00")

        # Ventas por día (para gráfico)
        from django.db.models.functions import TruncDate
        daily = (
            qs.annotate(day=TruncDate("sale_date"))
              .values("day")
              .annotate(total=Sum("total"), count=Count("id_sale"))
              .order_by("day")
        )

        # Productos más vendidos (top 15 por cantidad)
        top_products = (
            SaleDetail.objects
            .filter(sale__in=qs)
            .values("product__name", "product__category__name")
            .annotate(qty=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-qty")[:15]
        )

        # Top por categoría (top 10 por cantidad)
        top_categories = (
            SaleDetail.objects
            .filter(sale__in=qs)
            .values("product__category__name")
            .annotate(qty=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")[:10]
        )

        # Ventas por método de pago
        by_payment = (
            qs.values("payment__name")
              .annotate(count=Count("id_sale"), total=Sum("total"))
              .order_by("-total")
        )

        # Top vendedores (top 5 por ingresos)
        top_sellers = (
            qs.values("seller__name", "seller__last_name")
              .annotate(count=Count("id_sale"), total=Sum("total"))
              .order_by("-total")[:5]
        )

        # Top clientes (top 5 por ingresos)
        top_customers = (
            qs.values("customer__name", "customer__last_name")
              .annotate(count=Count("id_sale"), total=Sum("total"))
              .order_by("-total")[:5]
        )

        # Selectores para filtros
        sellers   = Seller.objects.all().order_by("name")
        customers = Customer.objects.all().order_by("name")

        context.update({
            "title": "Reportes y Estadísticas",
            # filtros activos
            "date_from":   date_from.isoformat(),
            "date_to":     date_to.isoformat(),
            "sel_payment": request.GET.get("payment", ""),
            "sel_seller":  request.GET.get("seller", ""),
            "sel_customer":request.GET.get("customer", ""),
            "sel_min":     request.GET.get("min_total", ""),
            "sel_max":     request.GET.get("max_total", ""),
            # KPIs
            "total_revenue":  total_revenue,
            "total_sales":    total_sales,
            "avg_ticket":     avg_ticket,
            "total_discount": total_discount,
            # gráficos / tablas
            "daily_data":      list(daily),
            "top_products":    list(top_products),
            "top_categories":  list(top_categories),
            "by_payment":      list(by_payment),
            "top_sellers":     list(top_sellers),
            "top_customers":   list(top_customers),
            # selectores
            "sellers":   sellers,
            "customers": customers,
        })
        return context


# Exportación a Excel

class ExportReportsExcelView(LoginRequiredMixin, View):
    login_url = "/security/login/"

    def get(self, request, *args, **kwargs):
        # Si viene el parámetro ?all=1, exporta todo sin filtros
        if request.GET.get("all") == "1":
           qs, date_from, date_to = _apply_filters(request, force_all=True)
        else:
           qs, date_from, date_to = _apply_filters(request)

        wb = Workbook()

        # Colores corporativos
        RED_DARK  = "7F1D1D"
        RED_MED   = "991B1B"
        RED_LIGHT = "FEE2E2"
        WHITE     = "FFFFFF"
        GRAY_DARK = "1F2937"
        GRAY_LIGHT = "F9FAFB"
        AMBER     = "F59E0B"
        GREEN     = "16A34A"
        BLUE      = "2563EB"
        CYAN      = "0EA5E9"
        PURPLE    = "7C3AED"

        thin = Side(style="thin", color="E5E7EB")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def header_font(size=11, bold=True, color=WHITE):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def data_font(size=10, bold=False, color=GRAY_DARK):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def fill(color):
            return PatternFill("solid", fgColor=color)

        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def right():
            return Alignment(horizontal="right", vertical="center")

        def _col_width(ws, col, width):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ══════════════════════════════════════════════════════════════════
        # HOJA 1 — RESUMEN EJECUTIVO
        # ══════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        ws1.sheet_view.showGridLines = False

        # Título principal
        ws1.merge_cells("A1:H1")
        ws1["A1"] = "🛒  MY SUPERMARKET — REPORTE DE VENTAS"
        ws1["A1"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
        ws1["A1"].fill = fill(RED_DARK)
        ws1["A1"].alignment = center()
        ws1.row_dimensions[1].height = 42

        ws1.merge_cells("A2:H2")
        ws1["A2"] = f"Período: {date_from.strftime('%d/%m/%Y')} — {date_to.strftime('%d/%m/%Y')}     |     Generado el {date.today().strftime('%d/%m/%Y')}"
        ws1["A2"].font = Font(name="Arial", size=10, color="FCA5A5")
        ws1["A2"].fill = fill(RED_MED)
        ws1["A2"].alignment = center()
        ws1.row_dimensions[2].height = 22

        ws1.row_dimensions[3].height = 10

        # KPIs
        agg = qs.aggregate(
            rev=Sum("total"), cnt=Count("id_sale"),
            disc=Sum("discount"), iva=Sum("iva")
        )
        total_revenue  = float(agg["rev"]  or 0)
        total_sales    = int(agg["cnt"]    or 0)
        total_discount = float(agg["disc"] or 0)
        total_iva      = float(agg["iva"]  or 0)
        avg_ticket     = (total_revenue / total_sales) if total_sales else 0.0

        kpis = [
            ("💰 Ingresos Totales", f"${total_revenue:,.2f}", GREEN),
            ("🧾 Ventas Realizadas", f"{total_sales:,}",       BLUE),
            ("📊 Ticket Promedio",   f"${avg_ticket:,.2f}",    CYAN),
            ("🏷️ Descuentos",        f"${total_discount:,.2f}", AMBER),
            ("📋 IVA Recaudado",     f"${total_iva:,.2f}",     PURPLE),
        ]

        ws1.merge_cells("A4:B4")
        ws1["A4"] = "INDICADORES CLAVE DE DESEMPEÑO"
        ws1["A4"].font = Font(name="Arial", size=9, bold=True, color="9CA3AF")
        ws1["A4"].fill = fill(GRAY_LIGHT)
        ws1["A4"].alignment = Alignment(horizontal="left", vertical="center")
        ws1.row_dimensions[4].height = 18

        for i, (label, value, color) in enumerate(kpis):
            row = 5 + i
            ws1.row_dimensions[row].height = 32
            ws1.merge_cells(f"A{row}:B{row}")
            ws1[f"A{row}"] = label
            ws1[f"A{row}"].font = Font(name="Arial", size=10, bold=True, color=color)
            ws1[f"A{row}"].fill = fill(GRAY_LIGHT)
            ws1[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws1[f"A{row}"].border = thin_border

            ws1[f"C{row}"] = value
            ws1[f"C{row}"].font = Font(name="Arial", size=13, bold=True, color=color)
            ws1[f"C{row}"].alignment = center()
            ws1[f"C{row}"].border = thin_border

        # ── Ventas por método de pago ─────────────────────────────────────
        row_pay = 4
        ws1.merge_cells("E4:H4")
        ws1["E4"] = "VENTAS POR MÉTODO DE PAGO"
        ws1["E4"].font = Font(name="Arial", size=9, bold=True, color="9CA3AF")
        ws1["E4"].fill = fill(GRAY_LIGHT)
        ws1["E4"].alignment = Alignment(horizontal="left", vertical="center")

        headers_pay = ["Método de Pago", "N° Ventas", "Total $", "% del Total"]
        for ci, h in enumerate(headers_pay, 5):
            c = ws1.cell(row=5, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(RED_DARK)
            c.alignment = center()
            c.border = thin_border

        by_payment = (
            qs.values("payment__name")
              .annotate(count=Count("id_sale"), total=Sum("total"))
              .order_by("-total")
        )

        for ri, p in enumerate(by_payment, 6):
            pct = (float(p["total"] or 0) / total_revenue * 100) if total_revenue else 0
            row_data = [
                p["payment__name"] or "—",
                p["count"],
                f"${float(p['total'] or 0):,.2f}",
                f"{pct:.1f}%",
            ]
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            for ci, val in enumerate(row_data, 5):
                c = ws1.cell(row=ri, column=ci, value=val)
                c.font = data_font()
                c.fill = fill(bg)
                c.alignment = center()
                c.border = thin_border
            ws1.row_dimensions[ri].height = 20

        # Anchos
        for col, w in [(1, 28), (2, 14), (3, 18), (4, 2), (5, 24), (6, 12), (7, 16), (8, 12)]:
            _col_width(ws1, col, w)

        # ══════════════════════════════════════════════════════════════════
        # HOJA 2 — DETALLE DE VENTAS
        # ══════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Detalle de Ventas")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:J1")
        ws2["A1"] = "DETALLE DE VENTAS"
        ws2["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
        ws2["A1"].fill = fill(RED_DARK)
        ws2["A1"].alignment = center()
        ws2.row_dimensions[1].height = 32

        headers_det = ["#","Fecha","Cliente","Vendedor","Método Pago",
                       "Subtotal $","IVA $","Descuento $","Total $","UUID"]
        for ci, h in enumerate(headers_det, 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(RED_MED)
            c.alignment = center()
            c.border = thin_border
        ws2.row_dimensions[2].height = 22

        sales_qs = qs.select_related("customer", "seller", "payment").order_by("-sale_date")
        for ri, s in enumerate(sales_qs, 3):
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            row_data = [
                s.id_sale,
                s.sale_date.strftime("%d/%m/%Y %H:%M") if s.sale_date else "",
                s.customer.get_full_name() if s.customer else "—",
                s.seller.get_full_name()   if s.seller   else "—",
                s.payment.name             if s.payment  else "—",
                float(s.subtotal or 0),
                float(s.iva or 0),
                float(s.discount or 0),
                float(s.total or 0),
                str(s.idempotency_key) if s.idempotency_key else "",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font = data_font()
                c.fill = fill(bg)
                c.border = thin_border
                if ci in (6, 7, 8, 9):
                    c.number_format = '"$"#,##0.00'
                    c.alignment = right()
                else:
                    c.alignment = Alignment(vertical="center", wrap_text=False)
            ws2.row_dimensions[ri].height = 18

        # Totales
        last = 2 + sales_qs.count()
        for ci, col_letter in [(6,"F"), (7,"G"), (8,"H"), (9,"I")]:
            c = ws2.cell(row=last+1, column=ci, value=f"=SUM({col_letter}3:{col_letter}{last})")
            c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            c.fill = fill(RED_DARK)
            c.number_format = '"$"#,##0.00'
            c.alignment = right()
            c.border = thin_border
        ws2.cell(row=last+1, column=1).value = "TOTAL"
        ws2.cell(row=last+1, column=1).font = Font(name="Arial", size=10, bold=True, color=WHITE)
        ws2.cell(row=last+1, column=1).fill = fill(RED_DARK)
        ws2.cell(row=last+1, column=1).alignment = center()

        col_ws2 = [5,18,24,20,18,13,11,13,13,36]
        for i, w in enumerate(col_ws2, 1):
            _col_width(ws2, i, w)

        # ══════════════════════════════════════════════════════════════════
        # HOJA 3 — PRODUCTOS MÁS VENDIDOS
        # ══════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Productos Más Vendidos")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:F1")
        ws3["A1"] = "PRODUCTOS MÁS VENDIDOS"
        ws3["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
        ws3["A1"].fill = fill(PURPLE)
        ws3["A1"].alignment = center()
        ws3.row_dimensions[1].height = 32

        headers_prod = ["Ranking","Producto","Categoría","Uds. Vendidas","Ingresos $","% Ing. Total"]
        for ci, h in enumerate(headers_prod, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(PURPLE)
            c.alignment = center()
            c.border = thin_border
        ws3.row_dimensions[2].height = 22

        top_products = (
            SaleDetail.objects
            .filter(sale__in=qs)
            .values("product__name", "product__category__name")
            .annotate(qty=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-qty")[:30]
        )
        total_rev_prod = sum(float(p["revenue"] or 0) for p in top_products)

        medal_colors = ["FFD700", "C0C0C0", "CD7F32"]
        for ri, p in enumerate(top_products, 3):
            bg = medal_colors[ri-3] if ri <= 5 else (GRAY_LIGHT if ri % 2 == 0 else WHITE)
            pct = (float(p["revenue"] or 0) / total_rev_prod * 100) if total_rev_prod else 0
            row_data = [
                ri - 2,
                p["product__name"] or "—",
                p["product__category__name"] or "Sin categoría",
                int(p["qty"] or 0),
                float(p["revenue"] or 0),
                f"{pct:.1f}%",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", size=10,
                              bold=(ri <= 5),
                              color=GRAY_DARK if ri > 5 else "000000")
                c.fill = fill(bg)
                c.alignment = center() if ci in (1, 3, 4, 6) else Alignment(vertical="center")
                c.border = thin_border
                if ci == 5:
                    c.number_format = '"$"#,##0.00'
                    c.alignment = right()
            ws3.row_dimensions[ri].height = 20

        col_ws3 = [10, 34, 22, 16, 16, 14]
        for i, w in enumerate(col_ws3, 1):
            _col_width(ws3, i, w)

        # ══════════════════════════════════════════════════════════════════
        # HOJA 4 — ANÁLISIS POR CATEGORÍA
        # ══════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Por Categoría")
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:E1")
        ws4["A1"] = "ANÁLISIS DE VENTAS POR CATEGORÍA"
        ws4["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
        ws4["A1"].fill = fill(GREEN)
        ws4["A1"].alignment = center()
        ws4.row_dimensions[1].height = 32

        headers_cat = ["Categoría","Uds. Vendidas","Ingresos $","% del Total","Ticket Prom. $"]
        for ci, h in enumerate(headers_cat, 1):
            c = ws4.cell(row=2, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(GREEN)
            c.alignment = center()
            c.border = thin_border
        ws4.row_dimensions[2].height = 22

        top_cat = (
            SaleDetail.objects
            .filter(sale__in=qs)
            .values("product__category__name")
            .annotate(qty=Sum("quantity"), revenue=Sum("subtotal"), items=Count("id_detail"))
            .order_by("-revenue")
        )
        total_rev_cat = sum(float(c2["revenue"] or 0) for c2 in top_cat)

        for ri, cat in enumerate(top_cat, 3):
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            pct = (float(cat["revenue"] or 0) / total_rev_cat * 100) if total_rev_cat else 0
            row_data = [
                cat["product__category__name"] or "Sin categoría",
                int(cat["qty"] or 0),
                float(cat["revenue"] or 0),
                f"{pct:.1f}%",
                float(cat["revenue"] or 0) / int(cat["items"] or 1),
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws4.cell(row=ri, column=ci, value=val)
                c.font = data_font()
                c.fill = fill(bg)
                c.alignment = center() if ci in (1, 2, 4) else right()
                c.border = thin_border
                if ci in (3, 5):
                    c.number_format = '"$"#,##0.00'
            ws4.row_dimensions[ri].height = 20

        for i, w in enumerate([28, 16, 16, 14, 16], 1):
            _col_width(ws4, i, w)

        # ══════════════════════════════════════════════════════════════════
        # HOJA 5 — TOP VENDEDORES Y CLIENTES
        # ══════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("Vendedores y Clientes")
        ws5.sheet_view.showGridLines = False

        ws5.merge_cells("A1:D1")
        ws5["A1"] = "TOP VENDEDORES"
        ws5["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
        ws5["A1"].fill = fill(BLUE)
        ws5["A1"].alignment = center()
        ws5.row_dimensions[1].height = 28

        headers_sell = ["Vendedor","N° Ventas","Total $","Ticket Prom. $"]
        for ci, h in enumerate(headers_sell, 1):
            c = ws5.cell(row=2, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(BLUE)
            c.alignment = center()
            c.border = thin_border
        ws5.row_dimensions[2].height = 20

        top_sellers = (
            qs.values("seller__name", "seller__last_name")
              .annotate(cnt=Count("id_sale"), total=Sum("total"))
              .order_by("-total")[:10]
        )
        for ri, s in enumerate(top_sellers, 3):
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            name = f"{s['seller__name'] or ''} {s['seller__last_name'] or ''}".strip() or "—"
            cnt   = int(s["cnt"] or 0)
            total = float(s["total"] or 0)
            avg   = (total / cnt) if cnt else 0.0
            row_data = [name, cnt, total, avg]
            for ci, val in enumerate(row_data, 1):
                c = ws5.cell(row=ri, column=ci, value=val)
                c.font = data_font()
                c.fill = fill(bg)
                c.alignment = center() if ci <= 2 else right()
                c.border = thin_border
                if ci in (3, 4):
                    c.number_format = '"$"#,##0.00'
            ws5.row_dimensions[ri].height = 20

        # Top Clientes (col F)
        offset = 6
        ws5.merge_cells(f"{get_column_letter(offset)}1:{get_column_letter(offset+3)}1")
        ws5.cell(row=1, column=offset).value = "TOP CLIENTES"
        ws5.cell(row=1, column=offset).font = Font(name="Arial", size=13, bold=True, color=WHITE)
        ws5.cell(row=1, column=offset).fill = fill(CYAN)
        ws5.cell(row=1, column=offset).alignment = center()

        headers_cust = ["Cliente","N° Compras","Total $","Ticket Prom. $"]
        for ci, h in enumerate(headers_cust, offset):
            c = ws5.cell(row=2, column=ci, value=h)
            c.font = header_font(9, True, WHITE)
            c.fill = fill(CYAN)
            c.alignment = center()
            c.border = thin_border

        top_customers = (
            qs.values("customer__name", "customer__last_name")
              .annotate(cnt=Count("id_sale"), total=Sum("total"))
              .order_by("-total")[:10]
        )
        for ri, cu in enumerate(top_customers, 3):
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            name = f"{cu['customer__name'] or ''} {cu['customer__last_name'] or ''}".strip() or "—"
            cnt   = int(cu["cnt"] or 0)
            total = float(cu["total"] or 0)
            avg   = (total / cnt) if cnt else 0.0
            row_data = [name, cnt, total, avg]
            for ci, val in enumerate(row_data, offset):
                c = ws5.cell(row=ri, column=ci, value=val)
                c.font = data_font()
                c.fill = fill(bg)
                c.alignment = center() if ci <= offset+1 else right()
                c.border = thin_border
                if ci in (offset+2, offset+3):
                    c.number_format = '"$"#,##0.00'
            ws5.row_dimensions[ri].height = 20

        for i, w in enumerate([24, 12, 14, 14, 4, 24, 12, 14, 14], 1):
            _col_width(ws5, i, w)

        # ══════════════════════════════════════════════════════════════════
        # Respuesta HTTP
        # ══════════════════════════════════════════════════════════════════
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"reporte_ventas_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response 